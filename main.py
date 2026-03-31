import pandas as pd
import glob
import os
import librosa
import unicodedata

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from feature_extraction import extract_features
from f0_estimation import compute_autocorrelation_f0
from classifier import classify_gender_from_f0


BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATASET_DIR = os.path.join(BASE_DIR, "dataset")
DATA_DIR = os.path.join(BASE_DIR, "data")
RESULTS_DIR = os.path.join(BASE_DIR, "results")


def adjust_excel_columns(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            try:
                cell_value = str(cell.value) if cell.value is not None else ""
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except Exception:
                pass

        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(file_path)


def load_metadata():
    excel_files = glob.glob(os.path.join(DATASET_DIR, "**", "*.xlsx"), recursive=True)

    if not excel_files:
        print("Hiç Excel dosyası bulunamadı.")
        return None

    df_list = []

    for file in excel_files:
        try:
            df = pd.read_excel(file)

            df["source_file"] = os.path.basename(file)
            df["source_folder"] = os.path.basename(os.path.dirname(file))

            if "Dosya_Adi" in df.columns:
                df = df[df["Dosya_Adi"] != "Dosya_Adi"]

            if "FILE NAME" in df.columns:
                df = df[df["FILE NAME"] != "FILE NAME"]

            if "File_Name" in df.columns:
                df = df[df["File_Name"] != "File_Name"]

            if "File name" in df.columns:
                df = df[df["File name"] != "File name"]

            if "File Name" in df.columns:
                df = df[df["File Name"] != "File Name"]

            if "                      FILE NAME" in df.columns:
                df = df[df["                      FILE NAME"] != "FILE NAME"]

            df_list.append(df)

        except Exception as e:
            print(f"Hata oluştu: {file} -> {e}")

    if not df_list:
        print("Okunabilen Excel dosyası yok.")
        return None

    master_df = pd.concat(df_list, ignore_index=True)
    return master_df


def save_metadata(master_df):
    os.makedirs(DATA_DIR, exist_ok=True)
    output_path = os.path.join(DATA_DIR, "master_metadata.xlsx")
    master_df.to_excel(output_path, index=False)
    adjust_excel_columns(output_path)
    print(f"Master metadata oluşturuldu: {output_path}")


def find_filename_column(df):
    possible_columns = [
        "Dosya_Adi",
        "FILE NAME",
        "File_Name",
        "File name",
        "File Name",
        "filename",
        "Filename",
        "                      FILE NAME"
    ]

    for col in possible_columns:
        if col in df.columns:
            return col

    return None


def get_value_from_possible_columns(row, possible_columns):
    for col in possible_columns:
        if col in row.index:
            value = row[col]
            if pd.notna(value):
                return value
    return None


def simplify_text(text):
    if text is None:
        return None

    text = str(text).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.replace("ı", "i")
    return text


def normalize_emotion(emotion):
    if emotion is None:
        return None

    emotion_simple = simplify_text(emotion)

    mapping = {
        "notr": "neutral",
        "notur": "neutral",
        "neutral": "neutral",
        "nötr": "neutral",
        "mutlu": "happy",
        "happy": "happy",
        "ofkeli": "angry",
        "okfeli": "angry",
        "öfkeli": "angry",
        "angry": "angry",
        "uzgun": "sad",
        "üzgun": "sad",
        "üzgün": "sad",
        "sad": "sad",
        "saskin": "surprised",
        "şaşkin": "surprised",
        "şaşkın": "surprised",
        "sasirma": "surprised",
        "surprised": "surprised",
        "surprise": "surprised"
    }

    return mapping.get(emotion_simple, emotion_simple)


def normalize_gender(gender):
    if gender is None:
        return None

    gender_simple = simplify_text(gender)

    mapping = {
        "e": "male",
        "erkek": "male",
        "male": "male",
        "k": "female",
        "kadin": "female",
        "kadın": "female",
        "female": "female",
        "c": "child",
        "cocuk": "child",
        "çocuk": "child",
        "child": "child"
    }

    return mapping.get(gender_simple, gender_simple)


def process_audio_files(df):
    filename_col = find_filename_column(df)

    if filename_col is None:
        print("Dosya adı kolonu bulunamadı.")
        print("Mevcut kolonlar:")
        print(df.columns)
        return None

    print(f"Kullanılan dosya adı kolonu: {filename_col}")

    found_count = 0
    missing_count = 0
    results = []

    for _, row in df.iterrows():
        try:
            file_name = row[filename_col]

            if pd.isna(file_name):
                continue

            file_name = str(file_name).strip()

            if not file_name.lower().endswith(".wav"):
                continue

            file_path = None
            for root, _, files in os.walk(DATASET_DIR):
                if file_name in files:
                    file_path = os.path.join(root, file_name)
                    break

            if file_path is None:
                missing_count += 1
                continue

            audio, sr = librosa.load(file_path, sr=None)
            energy, zcr, voiced_frames = extract_features(audio, sr)

            if energy is None or zcr is None or voiced_frames is None:
                continue

            avg_energy = float(energy.mean())
            avg_zcr = float(zcr.mean())
            voiced_frame_count = int(voiced_frames.sum())

            avg_f0 = compute_autocorrelation_f0(audio, sr)
            predicted_gender = classify_gender_from_f0(avg_f0)

            gender = get_value_from_possible_columns(
                row,
                ["Cinsiyet", "Gender", "  Gender"]
            )
            gender = normalize_gender(gender)

            age = get_value_from_possible_columns(
                row,
                ["Yas", "Age", "  Age"]
            )

            emotion = get_value_from_possible_columns(
                row,
                ["Duygu", "Feeling", "Feeling "]
            )
            emotion = normalize_emotion(emotion)

            results.append({
                "file_name": file_name,
                "sample_rate": sr,
                "num_samples": len(audio),
                "avg_energy": avg_energy,
                "avg_zcr": avg_zcr,
                "voiced_frame_count": voiced_frame_count,
                "avg_f0": avg_f0,
                "actual_gender": gender,
                "predicted_gender": predicted_gender,
                "age": age,
                "emotion": emotion
            })

            print(f"OK: {file_name}")
            found_count += 1

        except Exception as e:
            print(f"Hata: {e}")

    print("\nİşlem tamamlandı.")
    print(f"Bulunan dosya sayısı: {found_count}")
    print(f"Bulunamayan dosya sayısı: {missing_count}")

    return pd.DataFrame(results)


def save_feature_results(results_df):
    if results_df is None or results_df.empty:
        print("Kaydedilecek feature sonucu yok.")
        return

    os.makedirs(RESULTS_DIR, exist_ok=True)
    output_path = os.path.join(RESULTS_DIR, "features_summary.xlsx")
    results_df.to_excel(output_path, index=False)
    adjust_excel_columns(output_path)
    print(f"Feature sonuçları kaydedildi: {output_path}")


def print_performance_report(results_df):
    actual_col = "actual_gender"
    predict_col = "predicted_gender"
    f0_col = "avg_f0"

    valid_df = results_df[
        results_df[actual_col].isin(["male", "female", "child"]) &
        results_df[predict_col].isin(["male", "female", "child"])
    ].copy()

    print("\n" + "=" * 60)
    print("           SPEECH PROJECT - PERFORMANCE REPORT")
    print("=" * 60)

    conf_matrix = pd.crosstab(
        valid_df[actual_col],
        valid_df[predict_col]
    )

    conf_matrix = conf_matrix.reindex(
        index=["child", "female", "male"],
        columns=["child", "female", "male"],
        fill_value=0
    )

    conf_matrix["Total"] = conf_matrix.sum(axis=1)
    total_row = conf_matrix.sum(axis=0)
    conf_matrix.loc["Total"] = total_row
    conf_matrix.index.name = "Actual \\ Predicted"

    print("\nCONFUSION MATRIX")
    print(conf_matrix)

    cm_path = os.path.join(RESULTS_DIR, "confusion_matrix.xlsx")
    conf_matrix.to_excel(cm_path)
    adjust_excel_columns(cm_path)
    print(f"Confusion matrix kaydedildi: {cm_path}")

    performance_list = []

    for cls in ["male", "female", "child"]:
        subset = valid_df[valid_df[actual_col] == cls]

        if len(subset) == 0:
            continue

        n_samples = len(subset)
        avg_f = subset[f0_col].mean()
        std_f = subset[f0_col].std()
        correct = (subset[actual_col] == subset[predict_col]).sum()
        success = (correct / n_samples) * 100

        performance_list.append({
            "Class": cls.capitalize(),
            "Samples": n_samples,
            "Avg F0 (Hz)": round(avg_f, 2),
            "Std Dev": round(std_f, 2),
            "Success (%)": round(success, 2)
        })

    performance_df = pd.DataFrame(performance_list)

    print("\nCLASS-BASED PERFORMANCE")
    print(performance_df)

    perf_path = os.path.join(RESULTS_DIR, "performance_table.xlsx")
    performance_df.to_excel(perf_path, index=False)
    adjust_excel_columns(perf_path)
    print(f"Performance table kaydedildi: {perf_path}")

    overall_acc = (valid_df[actual_col] == valid_df[predict_col]).mean() * 100

    print("\n" + "*" * 60)
    print(f"OVERALL ACCURACY: %{overall_acc:.2f}")
    print("*" * 60)


if __name__ == "__main__":
    df = load_metadata()

    if df is not None:
        print("Toplam kayıt:", len(df))
        print(df.head())
        print(df.columns)

        save_metadata(df)
        results_df = process_audio_files(df)

        if results_df is not None and not results_df.empty:
            valid_df = results_df[
                results_df["actual_gender"].isin(["male", "female", "child"])
            ]

            if not valid_df.empty:
                accuracy = (
                    (valid_df["actual_gender"] == valid_df["predicted_gender"]).mean() * 100
                )
                print(f"Toplam Başarı Oranı: %{accuracy:.2f}")
            else:
                print("Accuracy hesaplamak için uygun etiket bulunamadı.")

        save_feature_results(results_df)

        if results_df is not None and not results_df.empty:
            print_performance_report(results_df)